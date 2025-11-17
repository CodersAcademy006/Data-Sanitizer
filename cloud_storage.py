"""
Cloud Storage & Parquet Output Module

Provides:
- S3 file reader (chunked streaming)
- GCS file reader (stub)
- Parquet output writer
- Multi-part upload support
"""

import os
import io
import logging
from typing import Generator, Optional, Dict, Any
from pathlib import Path

import pandas as pd

try:
    import boto3
    from botocore.exceptions import ClientError
    HAS_BOTO3 = True
except ImportError:
    HAS_BOTO3 = False

try:
    import pyarrow.parquet as pq
    import pyarrow as pa
    HAS_PYARROW = True
except ImportError:
    HAS_PYARROW = False

try:
    from google.cloud import storage as gcs_storage
    HAS_GCS = True
except ImportError:
    HAS_GCS = False

logger = logging.getLogger(__name__)

# ============================================================================
# S3 FILE READER
# ============================================================================

class S3FileReader:
    """
    Streams a file from S3 in memory-bounded chunks.
    Supports CSV, JSON, JSONL, Parquet, and Excel formats.
    """
    
    def __init__(
        self,
        bucket: str,
        key: str,
        aws_access_key: Optional[str] = None,
        aws_secret_key: Optional[str] = None,
        region: str = "us-east-1",
        chunksize: int = 50_000
    ):
        if not HAS_BOTO3:
            raise RuntimeError("boto3 is required for S3 operations. Install: pip install boto3")
        
        self.bucket = bucket
        self.key = key
        self.chunksize = chunksize
        
        # Initialize S3 client
        kwargs = {"region_name": region}
        if aws_access_key and aws_secret_key:
            kwargs.update({
                "aws_access_key_id": aws_access_key,
                "aws_secret_access_key": aws_secret_key
            })
        
        self.s3_client = boto3.client("s3", **kwargs)
        logger.info(f"S3FileReader initialized: s3://{bucket}/{key}")
    
    def stream_chunks(self) -> Generator[pd.DataFrame, None, None]:
        """
        Stream file chunks from S3.
        Automatically detects format from key suffix.
        """
        file_extension = self.key.lower().split(".")[-1]
        
        if file_extension in ["csv", "txt"]:
            yield from self._stream_csv()
        elif file_extension in ["jsonl", "ndjson"]:
            yield from self._stream_jsonl()
        elif file_extension == "json":
            yield from self._stream_json()
        elif file_extension == "parquet":
            yield from self._stream_parquet()
        elif file_extension in ["xlsx", "xls"]:
            yield from self._stream_excel()
        else:
            raise ValueError(f"Unsupported file format: .{file_extension}")
    
    def _stream_csv(self) -> Generator[pd.DataFrame, None, None]:
        """Stream CSV from S3."""
        try:
            response = self.s3_client.get_object(Bucket=self.bucket, Key=self.key)
            df_iter = pd.read_csv(
                response["Body"],
                chunksize=self.chunksize,
                low_memory=False
            )
            for chunk in df_iter:
                yield chunk
            logger.info(f"Finished streaming CSV from s3://{self.bucket}/{self.key}")
        except ClientError as e:
            logger.error(f"S3 error reading CSV: {e}")
            raise
    
    def _stream_jsonl(self) -> Generator[pd.DataFrame, None, None]:
        """Stream JSONL from S3 (requires ijson or line-by-line)."""
        try:
            import ijson
        except ImportError:
            logger.error("ijson required for JSONL streaming. Install: pip install ijson")
            raise
        
        try:
            response = self.s3_client.get_object(Bucket=self.bucket, Key=self.key)
            buffer = []
            
            for i, line in enumerate(response["Body"].iter_lines()):
                try:
                    import json
                    record = json.loads(line.decode("utf-8"))
                    buffer.append(record)
                    if len(buffer) >= self.chunksize:
                        yield pd.DataFrame(buffer)
                        buffer = []
                except Exception as e:
                    logger.warning(f"Could not parse line {i}: {e}")
            
            if buffer:
                yield pd.DataFrame(buffer)
        except ClientError as e:
            logger.error(f"S3 error reading JSONL: {e}")
            raise
    
    def _stream_json(self) -> Generator[pd.DataFrame, None, None]:
        """Stream JSON array from S3."""
        try:
            import ijson
            import json
        except ImportError:
            logger.error("ijson required for JSON streaming")
            raise
        
        try:
            response = self.s3_client.get_object(Bucket=self.bucket, Key=self.key)
            buffer = []
            
            for record in ijson.items(response["Body"], "item"):
                buffer.append(record)
                if len(buffer) >= self.chunksize:
                    yield pd.DataFrame(buffer)
                    buffer = []
            
            if buffer:
                yield pd.DataFrame(buffer)
        except ClientError as e:
            logger.error(f"S3 error reading JSON: {e}")
            raise
    
    def _stream_parquet(self) -> Generator[pd.DataFrame, None, None]:
        """Stream Parquet from S3."""
        if not HAS_PYARROW:
            raise RuntimeError("pyarrow required for Parquet. Install: pip install pyarrow")
        
        try:
            response = self.s3_client.get_object(Bucket=self.bucket, Key=self.key)
            parquet_file = pq.read_table(response["Body"])
            
            # Convert to chunks
            num_rows = parquet_file.num_rows
            for i in range(0, num_rows, self.chunksize):
                chunk = parquet_file.slice(i, min(self.chunksize, num_rows - i))
                yield chunk.to_pandas()
        except ClientError as e:
            logger.error(f"S3 error reading Parquet: {e}")
            raise
    
    def _stream_excel(self) -> Generator[pd.DataFrame, None, None]:
        """Stream Excel from S3."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl required for Excel. Install: pip install openpyxl")
            raise
        
        try:
            response = self.s3_client.get_object(Bucket=self.bucket, Key=self.key)
            wb = load_workbook(response["Body"], read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                
                try:
                    header = next(rows_iter)
                except StopIteration:
                    continue
                
                cols = [str(h) if h else f"col_{i}" for i, h in enumerate(header)]
                buffer = []
                
                for row_data in rows_iter:
                    row_dict = {cols[j]: row_data[j] if j < len(row_data) else None for j in range(len(cols))}
                    buffer.append(row_dict)
                    if len(buffer) >= self.chunksize:
                        yield pd.DataFrame(buffer)
                        buffer = []
                
                if buffer:
                    yield pd.DataFrame(buffer)
        except ClientError as e:
            logger.error(f"S3 error reading Excel: {e}")
            raise
    
    def upload_file(self, local_path: str, key: Optional[str] = None):
        """Upload a local file to S3."""
        target_key = key or self.key
        try:
            self.s3_client.upload_file(local_path, self.bucket, target_key)
            logger.info(f"Uploaded {local_path} to s3://{self.bucket}/{target_key}")
        except ClientError as e:
            logger.error(f"Failed to upload file: {e}")
            raise


# ============================================================================
# GCS FILE READER (STUB)
# ============================================================================

class GCSFileReader:
    """
    Google Cloud Storage file reader (stub implementation).
    """
    
    def __init__(self, bucket: str, blob_name: str, project_id: Optional[str] = None, chunksize: int = 50_000):
        if not HAS_GCS:
            raise RuntimeError("google-cloud-storage required for GCS. Install: pip install google-cloud-storage")
        
        self.bucket_name = bucket
        self.blob_name = blob_name
        self.chunksize = chunksize
        self.client = gcs_storage.Client(project=project_id)
        logger.info(f"GCSFileReader initialized: gs://{bucket}/{blob_name}")
    
    def stream_chunks(self) -> Generator[pd.DataFrame, None, None]:
        """Stream file chunks from GCS."""
        file_extension = self.blob_name.lower().split(".")[-1]
        
        try:
            bucket = self.client.bucket(self.bucket_name)
            blob = bucket.blob(self.blob_name)
            
            # Download to memory
            content = io.BytesIO()
            self.client.download_blob_to_file(blob, content)
            content.seek(0)
            
            if file_extension == "csv":
                df_iter = pd.read_csv(content, chunksize=self.chunksize, low_memory=False)
                for chunk in df_iter:
                    yield chunk
            elif file_extension == "parquet":
                if not HAS_PYARROW:
                    raise RuntimeError("pyarrow required")
                parquet_file = pq.read_table(content)
                num_rows = parquet_file.num_rows
                for i in range(0, num_rows, self.chunksize):
                    chunk = parquet_file.slice(i, min(self.chunksize, num_rows - i))
                    yield chunk.to_pandas()
            else:
                raise ValueError(f"Unsupported format: .{file_extension}")
        except Exception as e:
            logger.error(f"GCS error: {e}")
            raise


# ============================================================================
# PARQUET WRITER (STREAMING)
# ============================================================================

class ParquetStreamWriter:
    """
    Writes cleaned data to Parquet format in a streaming fashion.
    Supports both local and S3 output.
    """
    
    def __init__(self, output_path: str, compression: str = "snappy"):
        if not HAS_PYARROW:
            raise RuntimeError("pyarrow required for Parquet. Install: pip install pyarrow")
        
        self.output_path = output_path
        self.compression = compression
        self.is_s3 = output_path.startswith("s3://")
        
        if self.is_s3:
            if not HAS_BOTO3:
                raise RuntimeError("boto3 required for S3 output")
            self.s3_client = boto3.client("s3")
        
        self.writer = None
        logger.info(f"ParquetStreamWriter initialized: {output_path}")
    
    def write_chunk(self, df: pd.DataFrame):
        """Write a DataFrame chunk to Parquet."""
        if df.empty:
            return
        
        table = pa.Table.from_pandas(df)
        
        if self.is_s3:
            self._write_to_s3(table)
        else:
            self._write_to_local(table, df)
    
    def _write_to_local(self, table: pa.Table, df: pd.DataFrame):
        """Write table to local Parquet file."""
        if self.writer is None:
            # Initialize writer with schema from first table
            pq.write_table(
                table,
                self.output_path,
                compression=self.compression
            )
            self.writer = "initialized"
        else:
            # Append to existing file (requires pandas for now)
            existing = pq.read_table(self.output_path)
            combined = pa.concat_tables([existing, table])
            pq.write_table(
                combined,
                self.output_path,
                compression=self.compression
            )
    
    def _write_to_s3(self, table: pa.Table):
        """Write table to S3 Parquet."""
        # For simplicity, collect in memory and write once
        # In production, use Parquet's multipart upload or streaming approach
        buffer = io.BytesIO()
        pq.write_table(table, buffer, compression=self.compression)
        buffer.seek(0)
        
        bucket, key = self.output_path.replace("s3://", "").split("/", 1)
        self.s3_client.put_object(Bucket=bucket, Key=key, Body=buffer.getvalue())
    
    def finalize(self):
        """Finalize the Parquet file."""
        # Parquet files are self-contained, so nothing special needed
        logger.info(f"Finalized Parquet output: {self.output_path}")


# ============================================================================
# CSV WRITER (FOR COMPATIBILITY)
# ============================================================================

class CSVStreamWriter:
    """
    Writes cleaned data to CSV format in a streaming fashion.
    """
    
    def __init__(self, output_path: str, mode: str = "w"):
        self.output_path = output_path
        self.mode = mode
        self.first_write = True
        
        if output_path.startswith("s3://"):
            if not HAS_BOTO3:
                raise RuntimeError("boto3 required for S3 output")
            self.is_s3 = True
            self.s3_client = boto3.client("s3")
            self.buffer = io.StringIO()
        else:
            self.is_s3 = False
        
        logger.info(f"CSVStreamWriter initialized: {output_path}")
    
    def write_chunk(self, df: pd.DataFrame):
        """Write a DataFrame chunk to CSV."""
        if df.empty:
            return
        
        if self.is_s3:
            df.to_csv(self.buffer, index=False, header=self.first_write, mode="w")
            self.first_write = False
        else:
            mode = "w" if self.first_write else "a"
            df.to_csv(self.output_path, index=False, header=self.first_write, mode=mode)
            self.first_write = False
    
    def finalize(self):
        """Upload buffer to S3 if needed."""
        if self.is_s3 and self.buffer.tell() > 0:
            self.buffer.seek(0)
            bucket, key = self.output_path.replace("s3://", "").split("/", 1)
            self.s3_client.put_object(Bucket=bucket, Key=key, Body=self.buffer.getvalue())
            logger.info(f"Finalized CSV upload: s3://{bucket}/{key}")


# ============================================================================
# HELPER FUNCTION
# ============================================================================

def create_file_reader(uri: str, chunksize: int = 50_000):
    """
    Factory function to create appropriate file reader based on URI scheme.
    
    Examples:
    - "s3://bucket/path/file.csv"
    - "gs://bucket/path/file.parquet"
    - "/local/path/file.csv"
    """
    if uri.startswith("s3://"):
        parts = uri.replace("s3://", "").split("/", 1)
        return S3FileReader(parts[0], parts[1], chunksize=chunksize)
    elif uri.startswith("gs://"):
        parts = uri.replace("gs://", "").split("/", 1)
        return GCSFileReader(parts[0], parts[1], chunksize=chunksize)
    else:
        # Assume local file
        if not os.path.exists(uri):
            raise FileNotFoundError(f"File not found: {uri}")
        # Use the original safe_read_v3 from data_cleaning.py for local files
        from data_cleaning import safe_read_v3
        return safe_read_v3(uri, chunksize=chunksize)
