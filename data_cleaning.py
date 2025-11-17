# ----------------------------------------------------------------------------
# SECTION 1: SETUP (IMPORTS & INSTALLS)
# ----------------------------------------------------------------------------

# Install necessary libraries for streaming Excel and JSON
import os
import sqlite3
import hashlib
import itertools
import pandas as pd
import numpy as np
import io
import warnings
import json
import re
import logging
import math
import heapq
import struct
from collections import defaultdict, Counter
from difflib import get_close_matches, SequenceMatcher # <-- BUG FIX: Imported SequenceMatcher
# Attempt to import Colab file uploader; fall back gracefully if not available.
try:
    from google.colab import files  # type: ignore[import]  # <-- FEATURE: Added for file upload
    _HAS_COLAB = True
except Exception:
    files = None
    _HAS_COLAB = False
    warnings.warn("google.colab.files not available in this environment; upload option will be disabled.")

# Optional: ijson can be finicky.
try:
    import ijson  # type: ignore[import]
    _HAS_IJSON = True
except Exception:
    _HAS_IJSON = False

try:
    from openpyxl import load_workbook  # type: ignore[import]
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


# Setup basic logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ----------------------------------------------------------------------------
# SECTION 2: CORE ALGORITHM - INGESTION (safe_read_v3)
# ----------------------------------------------------------------------------

def flatten_json(obj, parent_key="", sep="."):
    """
    Recursively flatten JSON-like dicts into a flat dict with dot keys.
    Lists are converted to JSON strings by default (so the schema is stable).
    This is deterministic.
    """
    items = {}
    if obj is None:
        return {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.update(flatten_json(v, new_key, sep=sep))
            elif isinstance(v, list):
                # convert lists to JSON strings to keep schema stable
                try:
                    items[new_key] = json.dumps(v, ensure_ascii=False)
                except Exception:
                    items[new_key] = str(v)
            else:
                items[new_key] = v
    else:
        # not a dict — return as-is under parent_key
        items[parent_key] = obj
    return items

def discover_json_array_schema(path, max_keys=100_000):
    """
    Stream the JSON array and collect the union of flattened keys across objects.
    Uses ijson (streaming) so memory usage = O(#unique_keys).
    """
    if not _HAS_IJSON:
        raise RuntimeError("ijson is required to stream large JSON arrays. Install: pip install ijson")
    keys = set()
    try:
        with open(path, "rb") as f:
            # iterate over items in top-level array: prefix 'item'
            for obj in ijson.items(f, "item"):
                flat = flatten_json(obj)
                for k in flat.keys():
                    keys.add(k)
                    if len(keys) >= max_keys:
                        warnings.warn(f"Reached max_keys={max_keys} while discovering JSON schema. Consider increasing limit.")
                        return list(sorted(keys))
    except ijson.common.IncompleteJSONError:
         # This can happen if the file is not a JSON array.
         # Let's try to read it as a single object.
         try:
            with open(path, "rb") as f:
                obj = json.load(f)
                flat = flatten_json(obj)
                keys.update(flat.keys())
         except Exception as e:
            logger.error(f"Failed to read JSON file {path} as either array or object. Error: {e}")
            return [] # Return empty schema
    return list(sorted(keys))

def discover_jsonl_schema(path, max_keys=100_000):
    """
    Deterministically scan a JSONL / NDJSON file line-by-line, flatten each object,
    and collect the union of flattened keys. Memory = O(#unique_keys).
    """
    keys = set()
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception as e:
                raise ValueError(f"Invalid JSON on line {i+1}: {e}")
            flat = flatten_json(obj)
            for k in flat.keys():
                keys.add(k)
                if len(keys) >= max_keys:
                    warnings.warn(f"Reached max_keys={max_keys} while discovering JSONL schema. Consider increasing limit.")
                    return list(sorted(keys))
    return list(sorted(keys))

def jsonl_generator_with_schema(path, schema_keys, chunksize=50000):
    """
    Stream JSONL file line-by-line, flatten objects, and yield DataFrame chunks that
    conform to schema_keys (missing keys => None).
    """
    buffer = []
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                # Handle case where a line might not be valid JSON
                continue
                
            flat = flatten_json(obj)
            # build dict aligned to schema_keys (fast comprehension)
            buffer.append({k: flat.get(k, None) for k in schema_keys})
            if len(buffer) >= chunksize:
                yield pd.DataFrame(buffer)
                buffer = []
    if buffer:
        yield pd.DataFrame(buffer)

def safe_read_v3(path_or_buffer, chunksize=50000, json_schema_max_keys=100_000):
    """
    Unified, streaming reader that yields pandas.DataFrame chunks with stable schema.
    Supports:
      - CSV / TXT (pandas.read_csv with chunksize)
      - NDJSON / JSONL (two-pass schema discovery)
      - JSON array (streams via ijson, does schema discovery pass)
      - Excel (.xlsx/.xls) with streaming via openpyxl (read_only mode)
    """
    path = str(path_or_buffer)
    lower = path.lower()

    if lower.endswith(".csv") or lower.endswith(".txt"):
        logger.info(f"Reading CSV file: {path}")
        return pd.read_csv(path, chunksize=chunksize, low_memory=False)

    if lower.endswith(".jsonl") or lower.endswith(".ndjson"):
        logger.info(f"Discovering JSONL schema (two-pass) for stable columns: {path}")
        schema_keys = discover_jsonl_schema(path, max_keys=json_schema_max_keys)
        logger.info(f"Discovered {len(schema_keys)} keys in JSONL schema. Starting stream.")
        return jsonl_generator_with_schema(path, schema_keys, chunksize=chunksize)

    if lower.endswith(".json"):
        if not _HAS_IJSON:
            raise RuntimeError("To handle large JSON array files you must install ijson (`pip install ijson`).")
        logger.info(f"Discovering JSON array schema (streaming): {path}")
        schema_keys = discover_json_array_schema(path, max_keys=json_schema_max_keys)
        logger.info(f"Discovered {len(schema_keys)} keys in JSON array schema. Starting stream.")
        
        def gen_chunks():
            cols = schema_keys
            buffer = []
            try:
                with open(path, "rb") as f:
                    for i, obj in enumerate(ijson.items(f, "item")):
                        flat = flatten_json(obj)
                        buffer.append({k: flat.get(k, None) for k in cols})
                        if len(buffer) >= chunksize:
                            yield pd.DataFrame(buffer)
                            buffer = []
                if buffer:
                    yield pd.DataFrame(buffer)
            except ijson.common.IncompleteJSONError:
                # Fallback for single large JSON object (not in an array)
                if not buffer:
                     try:
                        with open(path, "rb") as f:
                            obj = json.load(f)
                            flat = flatten_json(obj)
                            yield pd.DataFrame([{k: flat.get(k, None) for k in cols}])
                     except Exception:
                        logger.error(f"Failed to parse {path} as single JSON object.")
                        
            except Exception as e:
                logger.error(f"Failed during JSON array streaming: {e}")
                
        return gen_chunks()

    if lower.endswith(".xlsx"):
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for streaming Excel files. Install: pip install openpyxl")
        logger.info(f"Streaming Excel file (per-sheet): {path}")
        wb = load_workbook(filename=path, read_only=True, data_only=True)
    
    elif lower.endswith(".xls"):
        # Use xlrd for legacy .xls format, but first check if it's actually CSV
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                first_line = f.readline()
                if ',' in first_line:
                    logger.info(f"Detected CSV content in .xls file; treating as CSV: {path}")
                    return pd.read_csv(path, chunksize=chunksize, low_memory=False)
        except Exception:
            pass
        
        # Try to open as binary Excel file
        try:
            import xlrd  # type: ignore
            import xlrd.sheet  # type: ignore
        except ImportError:
            raise RuntimeError("xlrd is required for legacy .xls files. Install: pip install xlrd")
        logger.info(f"Streaming legacy Excel file (per-sheet) with xlrd: {path}")
        try:
            workbook = xlrd.open_workbook(path, on_demand=True)
        except Exception as e:
            logger.warning(f"Failed to open as xlrd binary format: {e}. Trying as CSV fallback.")
            return pd.read_csv(path, chunksize=chunksize, low_memory=False)
        
        def gen_sheets_chunks_xls():
            for sheetname in workbook.sheet_names():
                logger.info(f"Streaming sheet: {sheetname}")
                sheet = workbook.sheet_by_name(sheetname)
                buffer = []
                if sheet.nrows < 1:
                    continue
                header = [str(cell.value) if cell.value is not None else f"col_{i}" for i, cell in enumerate(sheet.row(0))]
                for row_idx in range(1, sheet.nrows):
                    row_values = [cell.value for cell in sheet.row(row_idx)]
                    rowd = {header[j]: row_values[j] if j < len(row_values) else None for j in range(len(header))}
                    buffer.append(rowd)
                    if len(buffer) >= chunksize:
                        yield pd.DataFrame(buffer)
                        buffer = []
                if buffer:
                    yield pd.DataFrame(buffer)
        
        return gen_sheets_chunks_xls()
    
    else:
        # If we reach here and have xlsx/xls, use openpyxl
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for streaming Excel files. Install: pip install openpyxl")
        logger.info(f"Streaming Excel file (per-sheet): {path}")
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        
        def gen_sheets_chunks():
            for sheetname in wb.sheetnames:
                logger.info(f"Streaming sheet: {sheetname}")
                ws = wb[sheetname]
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header = next(rows_iter)
                except StopIteration:
                    continue # Empty sheet
                cols = [str(h) if h is not None else f"col_{i}" for i,h in enumerate(header)]
                buffer = []
                for i, row in enumerate(rows_iter):
                    rowd = {cols[j]: row[j] if j < len(row) else None for j in range(len(cols))}
                    buffer.append(rowd)
                    if len(buffer) >= chunksize:
                        yield pd.DataFrame(buffer)
                        buffer = []
                if buffer:
                    yield pd.DataFrame(buffer)
        return gen_sheets_chunks()

    # fallback to csv with explicit error
    try:
        logger.info(f"Attempting fallback to CSV for: {path}")
        return pd.read_csv(path, chunksize=chunksize, low_memory=False)
    except Exception as e:
        raise ValueError(f"Unsupported or unreadable file type: {path}. Error: {e}")

# ----------------------------------------------------------------------------
# SECTION 3: CORE ALGORITHM - STATISTICAL HELPERS
# ----------------------------------------------------------------------------

class DeterministicReservoir:
    """
    Keeps a deterministic, bounded sample of (key, value) pairs.
    Uses a hash(key_seed, row_id) as a deterministic priority. Keeps the smallest priorities.
    """
    def __init__(self, capacity=100000, salt="reservoir_v1"):
        self.capacity = int(capacity)
        self.salt = str(salt)
        self._heap = []  # max-heap by storing (-priority, row_id, value)
        
    def _priority(self, row_id):
        # deterministic 64-bit integer derived from row_id + salt
        h = hashlib.sha256(f"{self.salt}|{row_id}".encode('utf-8')).digest()
        # take first 8 bytes as unsigned int
        return struct.unpack(">Q", h[:8])[0]
        
    def add(self, row_id, value):
        p = self._priority(row_id)
        if len(self._heap) < self.capacity:
            heapq.heappush(self._heap, (-p, row_id, value))
        else:
            # check largest (heap root) which stores -p
            if p < -self._heap[0][0]:
                heapq.heapreplace(self._heap, (-p, row_id, value))
                
    def get_values(self):
        return [item[2] for item in self._heap]
        
    def get_priorities(self):
        return [(-item[0], item[1]) for item in self._heap]

def _shingles(text, k=5):
    """Generates a set of k-shingles from a text string."""
    text = str(text).lower()
    return set(text[i:i+k] for i in range(len(text) - k + 1))

def compute_minhash_signature(shingles, num_hashes=64):
    """
    Computes a MinHash signature from a set of shingles.
    This is a simple, self-contained implementation.
    """
    signature = []
    # Use different salts (seeds) for each hash function
    seeds = range(num_hashes)
    for seed in seeds:
        min_hash = float('inf')
        for shingle in shingles:
            # Simple hash: hash(shingle + seed)
            h = hashlib.sha256(f"{shingle}|{seed}".encode('utf-8')).digest()
            val = struct.unpack(">Q", h[:8])[0]
            if val < min_hash:
                min_hash = val
        
        # *** BUG FIX ***
        # Handle empty shingles by appending 0 instead of inf
        if min_hash == float('inf'):
            signature.append(0)
        else:
            signature.append(min_hash)
    return signature

def lsh_buckets_from_signature(signature, bands=16):
    """
    Computes LSH bucket hashes from a MinHash signature.
    """
    if not signature:
        return []
    rows = len(signature) // bands
    if rows == 0:
        return [hashlib.md5(str(signature).encode('utf-8')).hexdigest()]

    buckets = []
    for i in range(bands):
        band = signature[i*rows:(i+1)*rows]
        # Hash the band to a single bucket
        band_str = str(band).encode('utf-8')
        bucket_hash = hashlib.md5(band_str).hexdigest()
        buckets.append(bucket_hash)
    return buckets

# ----------------------------------------------------------------------------
# SECTION 4: CORE ALGORITHM - CLEANING HELPERS
# ----------------------------------------------------------------------------

def _chunk_row_hashes_vectorized(df, exclude_suffix="_was_imputed"):
    """
    Vectorized computation of MD5 row-hashes for all rows in a DataFrame chunk.
    Returns list of (hash, row_index_in_chunk) pairs.
    """
    cols = [c for c in df.columns if not c.endswith(exclude_suffix)]
    if not cols:
        # Handle edge case of no columns, or only imputed columns
        if df.empty:
            return []
        rows_concat = df.astype(str).agg("|".join, axis=1)
    else:
        rows_concat = df[cols].astype(str).agg("|".join, axis=1)
        
    # compute md5 for series
    hashes = rows_concat.apply(lambda s: hashlib.md5(s.encode('utf-8', errors='ignore')).hexdigest())
    # return list of tuples (hash, original_row_index)
    return list(zip(hashes.tolist(), df.index.tolist()))

def detect_outliers(df):
    """
    Simple outlier detection using IQR. Flags rows as outliers.
    """
    report = {"outliers": {}}
    for col in df.select_dtypes(include=np.number).columns:
        if col.endswith("_was_imputed"):
            continue
        try:
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1
            if pd.isna(IQR) or IQR == 0:
                continue # Skip if no variance
            lower = Q1 - 1.5 * IQR
            upper = Q3 + 1.5 * IQR
            mask = (df[col] < lower) | (df[col] > upper)
            if mask.sum() > 0:
                df[f"{col}_is_outlier"] = mask
                report["outliers"][col] = int(mask.sum())
        except Exception:
            pass # Ignore errors on columns with no variance, etc.
    return df, report

def clean_columns(df):
    """Simple column cleaning: strips whitespace from object columns."""
    report = {}
    for col in df.select_dtypes(include=['object']).columns:
        if col.endswith("_was_imputed"):
            continue
        try:
            # Check if it's already string, if not, convert
            if not all(isinstance(x, str) for x in df[col].dropna()):
                 df[col] = df[col].astype(str)
                 
            df[col] = df[col].str.strip()
            report[col] = "stripped"
        except Exception:
            pass
    return df, report

def text_normalization(df, keep_punctuation=True):
    """Simple text normalization: lowercase and optional punctuation removal."""
    report = {}
    for col in df.select_dtypes(include=['object']).columns:
        if col.endswith("_was_imputed"):
            continue
        try:
            # Check if it's already string, if not, convert
            if not all(isinstance(x, str) for x in df[col].dropna()):
                 df[col] = df[col].astype(str)
                 
            df[col] = df[col].str.lower()
            if not keep_punctuation:
                df[col] = df[col].str.replace(r'[^\w\s]', '', regex=True)
            report[col] = "normalized"
        except Exception:
            pass
    return df, report

def build_category_alias_map(series, similarity_threshold=0.86, max_categories=500):
    """
    Builds a map to merge similar-looking categories using difflib.
    """
    if series.nunique() > max_categories:
        return {} # Too many unique values, skip
        
    unique_vals = [str(x) for x in series.dropna().unique()]
    alias_map = {}
    done = set()
    for val in unique_vals:
        if val in done:
            continue
        # Find close matches to this value
        matches = get_close_matches(val, unique_vals, n=10, cutoff=similarity_threshold)
        # Use the first match (often itself) as the canonical name
        canonical = matches[0]
        for m in matches:
            alias_map[m] = canonical
            done.add(m)
    return alias_map

# ----------------------------------------------------------------------------
# SECTION 5: CORE ALGORITHM - DATABASE HELPERS
# ----------------------------------------------------------------------------

def create_sqlite_conn(path=":memory:", pragmas=None):
    """Creates a fast, WAL-enabled SQLite connection."""
    conn = sqlite3.connect(path, isolation_level=None, check_same_thread=False)
    cur = conn.cursor()
    # recommended pragmas for speed (safe in local context)
    default_pragmas = {
        "journal_mode":"WAL",
        "synchronous":"NORMAL",
        "temp_store":"MEMORY",
        "locking_mode":"EXCLUSIVE"
    }
    if pragmas is None:
        pragmas = default_pragmas
    for k,v in pragmas.items():
        try:
            cur.execute(f"PRAGMA {k}={v};")
        except Exception:
            pass
    return conn

def init_sqlite_dbs(conn):
    """Initializes the tables for deduplication and LSH samples."""
    cur = conn.cursor()
    # exact dedupe table (primary key on hash ensures uniqueness)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS row_hashes (
      hash TEXT PRIMARY KEY,
      first_seen_row INTEGER
    );
    """)
    # LSH buckets table: bucket -> row_id -> snippet
    cur.execute("""
    CREATE TABLE IF NOT EXISTS lsh_samples (
      bucket_key TEXT,
      sampled_row_id INTEGER,
      snippet TEXT,
      PRIMARY KEY (bucket_key, sampled_row_id)
    );
    """)
    conn.commit()

# ----------------------------------------------------------------------------
# SECTION 6: CORE ALGORITHM - THE TWO PASSES
# ----------------------------------------------------------------------------

def compute_global_stats_reservoir_schema_aware(
    path,
    conn,
    chunksize=50000,
    numeric_sample_mod=10,
    categorical_sample_mod=20,
    lsh_sample_mod=50,
    original_sample_mod=10,
    numeric_capacity=100000,
    max_original_sample_rows=1000,
    numeric_vote_threshold=0.5  # if >50% sampled values looked numeric -> treat as numeric
):
    """
    Deterministic first pass that:
      - streams input using safe_read_v3
      - uses deterministic reservoir sampling per column
      - votes on column types across the *entire* file (handles schema drift)
      - stores LSH sampled snippets to sqlite (batched)
    """
    logger.info("Starting schema-aware first pass (reservoirs + sqlite)")
    reader = safe_read_v3(path, chunksize=chunksize)
    
    # We must read the first chunk to establish an initial column list
    try:
        first_chunk = next(iter(reader))
        # Chain the first_chunk back to the reader iterator
        reader = itertools.chain([first_chunk], reader)
        cols = list(first_chunk.columns)
    except StopIteration:
        logger.warning(f"File {path} appears to be empty.")
        first_chunk = pd.DataFrame()
        cols = []
    except Exception as e:
        logger.error(f"Could not read first chunk from {path}: {e}")
        return {} # Cannot proceed
        
    reservoirs = {}  # col -> DeterministicReservoir
    numeric_votes = defaultdict(int)  # col -> how many sampled rows looked numeric
    sampled_votes = defaultdict(int)  # col -> how many rows were considered in sampling
    categorical_counts = defaultdict(Counter)
    original_samples = []
    row_id = 0

    MINHASH_NUM = 64
    LSH_BANDS = 16
    SHINGLE_K = 5
    cur = conn.cursor()

    def _bulk_insert_lsh(rows_to_insert):
        if not rows_to_insert:
            return
        cur.executemany("INSERT OR IGNORE INTO lsh_samples(bucket_key, sampled_row_id, snippet) VALUES (?, ?, ?);", rows_to_insert)
        conn.commit()

    def _is_numeric_like(val):
        """
        Robust check for numeric-like values, handling 'nan' and 'None' strings.
        """
        if pd.isna(val):
            return False
        if isinstance(val, (int, float, np.number)):
            return True
        s = str(val).strip()
        
        # *** BUG FIX ***
        # Handle common non-numeric strings explicitly
        if s == "" or s.lower() == 'nan' or s.lower() == 'none':
            return False
            
        try:
            float(s)
            return True
        except Exception:
            return False
    
    # Process all chunks
    for chunk in reader:
        rows_to_insert = []
        
        # Ensure new columns from schema drift are added to our global list
        new_cols = [c for c in chunk.columns if c not in cols]
        if new_cols:
            cols.extend(new_cols)

        for _, row in chunk.iterrows():
            # numeric periodic sampling
            if (row_id % numeric_sample_mod) == 0:
                for c in cols:
                    v = row.get(c)
                    sampled_votes[c] += 1
                    if _is_numeric_like(v):
                        numeric_votes[c] += 1
                        if c not in reservoirs:
                            reservoirs[c] = DeterministicReservoir(capacity=numeric_capacity, salt=f"{c}_v1")
                        try:
                            reservoirs[c].add(row_id, float(v))
                        except (ValueError, TypeError):
                            pass # Failed to cast, not numeric
                        
            # categorical periodic sampling
            if (row_id % categorical_sample_mod) == 0:
                for c in cols:
                    v = row.get(c)
                    if pd.notna(v):
                        categorical_counts[c][str(v).strip().lower()] += 1
                        
            # LSH periodic sampling
            if (row_id % lsh_sample_mod) == 0:
                # Use first 3 columns as-is for snippet
                snippet = " ".join([str(row.get(c,"")) for c in cols[:3]])
                shingles = _shingles(snippet, k=SHINGLE_K)
                sig = compute_minhash_signature(shingles, num_hashes=MINHASH_NUM)
                bks = lsh_buckets_from_signature(sig, bands=LSH_BANDS)
                for b in bks:
                    rows_to_insert.append((str(b), int(row_id), snippet))
                    
            # original sample
            if len(original_samples) < max_original_sample_rows and (row_id % original_sample_mod) == 0:
                original_samples.append(row.to_dict())
                
            row_id += 1
            
        _bulk_insert_lsh(rows_to_insert)

    # Decide numeric columns by vote threshold
    numeric_cols_final = [c for c, s in sampled_votes.items() if s > 0 and (numeric_votes[c] / float(s)) >= numeric_vote_threshold]
    
    # Compute medians from reservoirs
    medians = {}
    for c in numeric_cols_final:
        r = reservoirs.get(c)
        if r:
            vals = r.get_values()
            if vals:
                medians[c] = float(np.median(vals))

    modes = {c: cnt.most_common(1)[0][0] for c,cnt in categorical_counts.items() if cnt}
    
    stats = {
        "medians": medians,
        "modes": modes,
        "numeric_cols": numeric_cols_final,
        "text_cols": [c for c in cols if c not in numeric_cols_final],
        "all_cols": cols,
        "original_sample_df": pd.DataFrame(original_samples) if original_samples else pd.DataFrame(),
        "original_row_count": row_id,
        "minhash_params": {"num_hashes": MINHASH_NUM, "bands": LSH_BANDS, "shingle_k": SHINGLE_K},
        "sqlite_conn": conn
    }
    conn.commit()
    logger.info("Schema-aware first pass done rows=%d, numeric_cols=%d", row_id, len(numeric_cols_final))
    return stats

def clean_with_sqlite_dedupe_batched(path, output_dir, stats, chunksize=50000,
                                     keep_punctuation=True, drop_outliers=False, drop_outlier_columns=None,
                                     near_dup_threshold=0.85, csv_stream_path=None):
    """
    Batched / vectorized cleaning pass that:
      - computes all row hashes per chunk in vectorized form
      - queries sqlite for existing hashes in one batched query
      - inserts new hashes via executemany in one batch
      - streams cleaned chunk to CSV to avoid memory concat
      - fetches LSH candidate snippets for all bucket keys for the chunk in one query
    """
    conn = stats["sqlite_conn"]
    cur = conn.cursor()
    init_sqlite_dbs(conn)

    medians = stats.get("medians", {})
    modes = stats.get("modes", {})
    numeric_cols = stats.get("numeric_cols", [])
    text_cols = stats.get("text_cols", [])
    all_cols = stats.get("all_cols", [])
    minhash_params = stats.get("minhash_params", {"num_hashes":64,"bands":16,"shingle_k":5})
    MINHASH_NUM = minhash_params["num_hashes"]
    LSH_BANDS = minhash_params["bands"]
    SHINGLE_K = minhash_params["shingle_k"]

    # Prepare CSV streaming
    if csv_stream_path is None:
        csv_stream_path = os.path.join(output_dir, "cleaned_data.csv")
    first_write = True
    
    # If file exists already from previous run, remove it (safety)
    if os.path.exists(csv_stream_path):
        try:
            os.remove(csv_stream_path)
        except Exception:
            pass

    cleaned_row_count = 0
    parts_reports = []
    reader = safe_read_v3(path, chunksize=chunksize)

    for chunk_idx, chunk in enumerate(reader):
        logger.info(f"Processing chunk {chunk_idx+1} (batched DB ops) ...")
        if chunk.empty:
            continue
            
        # Ensure chunk has all columns from the global schema
        for col in all_cols:
            if col not in chunk.columns:
                chunk[col] = None
        # And ensure it's in the correct order
        chunk = chunk[all_cols]
        
        df = chunk.copy()
        chunk_report = {"near_duplicates_found": 0, "outliers": {}, "columns_fixed": {}, "text_normalized": {}, "imputed_counts": {}}

        # compute row hashes vectorized for chunk
        hashed_pairs = _chunk_row_hashes_vectorized(df)  # list of (hash, local_idx)
        hashes = [h for h, idx in hashed_pairs]

        # batch SELECT to find which hashes already exist
        existing_hashes = set()
        BATCH = 999 # Max variables in SQLite is 999
        if hashes:
            for i in range(0, len(hashes), BATCH):
                batch = hashes[i:i+BATCH]
                q = "SELECT hash FROM row_hashes WHERE hash IN ({seq})".format(seq=",".join("?"*len(batch)))
                cur.execute(q, batch)
                rows = cur.fetchall()
                existing_hashes.update([r[0] for r in rows])

        # determine keep_mask booleans and prepare inserts for those not present
        to_insert = []
        keep_local_idxs = []
        for (h, local_idx) in hashed_pairs:
            if h in existing_hashes:
                continue
            else:
                keep_local_idxs.append(local_idx)
                # We need to make sure we don't add the same hash twice in one batch
                existing_hashes.add(h) 
                to_insert.append((h, int(cleaned_row_count + len(keep_local_idxs) - 1)))
        
        # Bulk insert new hashes (keep-first semantics)
        if to_insert:
            cur.executemany("INSERT OR IGNORE INTO row_hashes(hash, first_seen_row) VALUES (?, ?);", to_insert)
            conn.commit()

        # Filter df to kept rows efficiently
        if keep_local_idxs:
            df_kept = df.loc[keep_local_idxs].reset_index(drop=True)
        else:
            df_kept = pd.DataFrame(columns=df.columns)  # empty

        if df_kept.empty:
            logger.info(f"Chunk {chunk_idx+1} was all duplicates.")
            parts_reports.append(chunk_report)
            continue

        # --- Start cleaning on df_kept ---
        
        # impute using medians/modes
        for col in df_kept.columns:
            if col.endswith("_was_imputed") or col.endswith("_is_outlier"):
                continue
            na_mask = df_kept[col].isna()
            if na_mask.sum() == 0:
                continue
            
            fill = None
            if col in medians and col in numeric_cols:
                fill = medians[col]
            elif col in modes:
                fill = modes[col]
            else:
                # fallback local
                if col in numeric_cols:
                    try:
                        fill = df_kept[col].median()
                    except Exception:
                        pass
                elif df_kept[col].dtype == 'object':
                    fill = df_kept[col].mode().iloc[0] if not df_kept[col].mode().empty else None
            
            if fill is not None and pd.notna(fill):
                df_kept[col + "_was_imputed"] = na_mask
                df_kept[col] = df_kept[col].fillna(fill)
                if df_kept[col + "_was_imputed"].any():
                    chunk_report["imputed_counts"][col] = int(df_kept[col + "_was_imputed"].sum())

        # detect outliers per-chunk
        df_kept, outlier_report = detect_outliers(df_kept)
        chunk_report["outliers"] = outlier_report.get("outliers", {})

        # clean columns and normalize text
        df_kept, col_report = clean_columns(df_kept)
        df_kept, text_report = text_normalization(df_kept, keep_punctuation=keep_punctuation)
        chunk_report["columns_fixed"] = col_report
        chunk_report["text_normalized"] = text_report

        # categorical merging
        for c in df_kept.columns:
            if df_kept[c].dtype == object and not c.endswith("_was_imputed"):
                unique_count = df_kept[c].nunique(dropna=True)
                if 2 <= unique_count <= 500:
                    try:
                        alias_map = build_category_alias_map(df_kept[c], similarity_threshold=0.86)
                        if alias_map:
                            df_kept[c] = df_kept[c].apply(lambda x: alias_map.get(x, x))
                    except Exception:
                        pass

        # NEAR-DUP detection
        bucket_keys_needed = set()
        kept_snippets = []
        for idx, row in df_kept.iterrows():
            snippet = " ".join([str(row.get(c,"")) for c in all_cols[:3]])
            shingles = _shingles(snippet, k=SHINGLE_K)
            sig = compute_minhash_signature(shingles, num_hashes=MINHASH_NUM)
            buckets = lsh_buckets_from_signature(sig, bands=LSH_BANDS)
            bucket_keys_needed.update(buckets)
            kept_snippets.append((idx, snippet, buckets))

        # bulk query candidate snippets
        bucket_to_candidates = defaultdict(list)
        if bucket_keys_needed:
            BK_list = list(bucket_keys_needed)
            for i in range(0, len(BK_list), BATCH):
                sub = BK_list[i:i+BATCH]
                q = "SELECT bucket_key, sampled_row_id, snippet FROM lsh_samples WHERE bucket_key IN ({})".format(",".join("?"*len(sub)))
                cur.execute(q, tuple(sub))
                for bk, rid, snip in cur.fetchall():
                    bucket_to_candidates[bk].append((rid, snip))

        # compute similarities in memory
        near_dup_pairs = []
        for (local_idx, snippet, buckets) in kept_snippets:
            seen_cands = set()
            for b in buckets:
                for cand_id, cand_snip in bucket_to_candidates.get(b, []):
                    if cand_id in seen_cands:
                        continue
                    seen_cands.add(cand_id)
                    
                    try:
                        sim = SequenceMatcher(None, snippet, cand_snip).ratio()
                    except Exception:
                        sim = 0
                        
                    if sim >= near_dup_threshold:
                        near_dup_pairs.append({"row_index": int(local_idx), "candidate_id": int(cand_id), "similarity": float(sim)})
                        
        chunk_report["near_duplicates_found"] = len(near_dup_pairs)

        # STREAM cleaned chunk to CSV
        mode = "w" if first_write else "a"
        header = first_write
        df_kept.to_csv(csv_stream_path, index=False, mode=mode, header=header)
        
        first_write = False
        cleaned_row_count += len(df_kept)
        parts_reports.append(chunk_report)

    conn.commit()
    logger.info(f"Finished cleaning. Total rows kept: {cleaned_row_count}")
    return cleaned_row_count, parts_reports # Return count and reports

# ----------------------------------------------------------------------------
# SECTION 7: CORE ALGORITHM - REPORTING
# ----------------------------------------------------------------------------

def generate_report_two_pass_fixed(
    original_sample_df,
    original_row_count,
    cleaned_row_count,
    after_sample_df, # A small sample from the cleaned file
    parts_reports,
    stats
):
    """
    Generates the final JSON report by aggregating chunk reports.
    """
    logger.info("Generating final cleaning report...")
    
    # Aggregate chunk-level reports
    total_near_dups = 0
    total_imputed = defaultdict(int)
    total_outliers = defaultdict(int)
    
    for report in parts_reports:
        total_near_dups += report.get("near_duplicates_found", 0)
        for col, count in report.get("imputed_counts", {}).items():
            total_imputed[col] += count
        for col, count in report.get("outliers", {}).items():
            total_outliers[col] += count
            
    # Calculate deduplication
    rows_dropped = original_row_count - cleaned_row_count
    
    # Get schema samples
    try:
        original_schema = {col: str(dtype) for col, dtype in original_sample_df.dtypes.items()}
    except Exception:
        original_schema = {}
        
    try:
        cleaned_schema = {col: str(dtype) for col, dtype in after_sample_df.dtypes.items()}
    except Exception:
        cleaned_schema = {}

    report_data = {
        "summary": {
            "original_row_count": original_row_count,
            "cleaned_row_count": cleaned_row_count,
            "rows_dropped_total": rows_dropped,
            "near_duplicates_flagged": total_near_dups,
        },
        "schema": {
            "original_schema_sample": original_schema,
            "cleaned_schema_sample": cleaned_schema,
            "numeric_cols_detected": stats.get("numeric_cols", []),
            "text_cols_detected": stats.get("text_cols", []),
        },
        "cleaning_details": {
            "imputed_counts": dict(total_imputed),
            "outliers_flagged": dict(total_outliers),
        },
        "global_stats_used": {
            "medians": stats.get("medians", {}),
            "modes": stats.get("modes", {}),
        },
        "samples": {
            "before_sample": original_sample_df.to_dict('records') if not original_sample_df.empty else [],
            "after_sample": after_sample_df.to_dict('records') if not after_sample_df.empty else [],
        }
    }
    return report_data

# ----------------------------------------------------------------------------
# SECTION 8: THE ORCHESTRATOR
# ----------------------------------------------------------------------------

def run_full_cleaning_pipeline_two_pass_sqlite_batched(
    path,
    output_dir=".",
    sqlite_path="cleaner_state.db",
    chunksize=50000,
    numeric_sample_mod=10,
    categorical_sample_mod=20,
    lsh_sample_mod=50,
    original_sample_mod=10,
    numeric_capacity=100000,
    max_original_sample_rows=1000,
    keep_punctuation=True,
    drop_outliers=False,
    drop_outlier_columns=None,
    near_dup_threshold=0.85
):
    """
    Orchestrator: removes existing sqlite_path (clean state), runs pass1 (batched),
    pass2 (batched), generates report.
    """
    # remove previous DB to guarantee isolated run
    if os.path.exists(sqlite_path):
        try:
            os.remove(sqlite_path)
            logger.info(f"Removed existing sqlite state file: {sqlite_path}")
        except Exception as e:
            logger.warning(f"Could not remove sqlite file {sqlite_path}: {e}")

    os.makedirs(output_dir, exist_ok=True)
    conn = create_sqlite_conn(sqlite_path)
    init_sqlite_dbs(conn)

    # Pass 1: reservoir + LSH inserted in bulk per chunk
    stats = compute_global_stats_reservoir_schema_aware(
        path, conn,
        chunksize=chunksize,
        numeric_sample_mod=numeric_sample_mod,
        categorical_sample_mod=categorical_sample_mod,
        lsh_sample_mod=lsh_sample_mod,
        original_sample_mod=original_sample_mod,
        numeric_capacity=numeric_capacity,
        max_original_sample_rows=max_original_sample_rows
    )
    
    if not stats:
        logger.error("First pass failed or file was empty. Aborting.")
        conn.close()
        return None, None

    stats['sqlite_conn'] = conn

    original_sample_df = stats.get("original_sample_df", pd.DataFrame())
    original_row_count = stats.get("original_row_count", 0)

    # Pass 2: batched cleaning and streaming to CSV
    cleaned_path = os.path.join(output_dir, "cleaned_data.csv")
    cleaned_row_count, parts_reports = clean_with_sqlite_dedupe_batched(
        path, output_dir, stats, chunksize=chunksize,
        keep_punctuation=keep_punctuation,
        drop_outliers=drop_outliers,
        drop_outlier_columns=drop_outlier_columns,
        near_dup_threshold=near_dup_threshold,
        csv_stream_path=cleaned_path
    )

    # Since we streamed cleaned output, load a small sample as 'after_sample'
    try:
        after_sample_df = pd.read_csv(cleaned_path, nrows=5)
    except Exception:
        after_sample_df = pd.DataFrame()

    # Build report
    report = generate_report_two_pass_fixed(
        original_sample_df,
        original_row_count,
        cleaned_row_count,
        after_sample_df,
        parts_reports,
        stats
    )

    # save the report JSON
    report_path = os.path.join(output_dir, "cleaning_report.json")
    try:
        with open(report_path, "w", encoding="utf-8") as f:
            # Use default=str to handle numpy types
            json.dump(report, f, indent=2, default=str)
        logger.info(f"Saved cleaning report to {report_path}")
    except Exception as e:
        logger.error(f"Failed to save cleaning report: {e}")

    # close DB connection
    try:
        conn.close()
    except:
        pass

    return cleaned_path, report_path


# ----------------------------------------------------------------------------
# SECTION 9: INTERACTIVE RUNNER (UPLOAD, PATH, OR DEMO)
# ----------------------------------------------------------------------------

def run_pipeline_on_user_file(file_path, chunksize=100000):
    """Helper function to run the full pipeline on a specified file."""
    try:
        # Define output paths
        output_dir = f"{os.path.basename(file_path)}_output"
        db_path = f"{os.path.basename(file_path)}_cleaner.db"
        
        logger.info(f"Starting pipeline for: {file_path}")
        logger.info(f"Output will be in: ./{output_dir}/")
        logger.info(f"Database will be at: ./{db_path}")

        # Run the full pipeline
        cleaned_path, report_path = run_full_cleaning_pipeline_two_pass_sqlite_batched(
            path=file_path,
            output_dir=output_dir,
            sqlite_path=db_path,
            chunksize=chunksize 
        )

        # --- Print results ---
        if cleaned_path and report_path:
            logger.info("="*30)
            logger.info(f"CLEANING RESULTS FOR: {file_path}")
            logger.info("="*30)
            with open(report_path, 'r') as f:
                print(json.dumps(json.load(f)['summary'], indent=2))
            
            print(f"\n--- Cleaned Data (Head) saved to {cleaned_path} ---")
            print(pd.read_csv(cleaned_path, nrows=5).head())
            
            print("\nTo download your cleaned file, run this in a new cell:")
            print(f"from google.colab import files\nfiles.download('{cleaned_path}')")
            
            print("\nTo download your report, run this in a new cell:")
            print(f"from google.colab import files\nfiles.download('{report_path}')")
        else:
             logger.error("Pipeline run failed for the specified file.")

    except Exception as e:
        logger.error(f"An error occurred during the pipeline run: {e}")
    


def run_demo_examples():
    """Helper function to create and run the built-in demo files."""
    logger.info("="*30)
    logger.info("OPTION 3: RUNNING BUILT-IN EXAMPLES")
    logger.info("="*30)

    logger.info("Setting up example data files...")

    # --- 1. Create a messy CSV file ---
    csv_data = """id,name,price,category,description
1,Product A,100,electronics,Good product
2,Product B,150.5,clothing,A duplicate row
2,Product B,150.5,clothing,A duplicate row
3,Product C,,electronics,Missing price
4,Product D,200,tools,
5,product a,101,Electronics,Similar to product A (near-dup)
6,Product F,99999,clothing,This is an outlier price
7,Product G,50,tools,Another tool
8,Product H,NaN,clothing,Another missing price
"""
    with open("test_data.csv", "w") as f:
        f.write(csv_data)

    # --- 2. Create a messy JSONL (ndjson) file with schema drift ---
    jsonl_data = """{"id": 10, "user": "alice", "value": 100}
{"id": 11, "user": "bob", "location": "NY"}
{"id": 10, "user": "alice", "value": 100}
{"id": 12, "user": "charlie", "value": null, "tags": ["a", "b"]}
"""
    with open("test_data.jsonl", "w") as f:
        f.write(jsonl_data)

    # --- 3. Create a messy Excel (xlsx) file ---
    try:
        excel_df = pd.DataFrame([
            {"item": "Pen", "stock": 100, "color": "Blue"},
            {"item": "Pen", "stock": 100, "color": "Blue"}, # duplicate
            {"item": "Pencil", "stock": 200, "color": "Yellow"},
            {"item": "Eraser", "stock": None, "color": "Pink"} # missing
        ])
        excel_df.to_excel("test_data.xlsx", sheet_name="Sheet1", index=False)
        _EXAMPLE_EXCEL_CREATED = True
    except Exception as e:
        _EXAMPLE_EXCEL_CREATED = False
        logger.error(f"Could not create Excel file (openpyxl writer might be missing): {e}")


    logger.info("--- 1. RUNNING PIPELINE ON CSV DATA ---")
    csv_cleaned_path, csv_report_path = run_full_cleaning_pipeline_two_pass_sqlite_batched(
        path="test_data.csv",
        output_dir="csv_output",
        sqlite_path="csv_cleaner.db",
        chunksize=5 # Use tiny chunksize for testing
    )

    logger.info("--- 2. RUNNING PIPELINE ON JSONL DATA ---")
    jsonl_cleaned_path, jsonl_report_path = run_full_cleaning_pipeline_two_pass_sqlite_batched(
        path="test_data.jsonl",
        output_dir="jsonl_output",
        sqlite_path="jsonl_cleaner.db",
        chunksize=2 # Use tiny chunksize for testing
    )

    if _HAS_OPENPYXL and _EXAMPLE_EXCEL_CREATED:
        logger.info("--- 3. RUNNING PIPELINE ON EXCEL DATA ---")
        excel_cleaned_path, excel_report_path = run_full_cleaning_pipeline_two_pass_sqlite_batched(
            path="test_data.xlsx",
            output_dir="excel_output",
            sqlite_path="excel_cleaner.db",
            chunksize=2 # Use tiny chunksize for testing
        )


    # --- 4. Print results ---
    logger.info("="*30)
    logger.info("CSV CLEANING RESULTS (EXAMPLE 2)")
    logger.info("="*30)
    try:
        with open(csv_report_path, 'r') as f:
            print(json.dumps(json.load(f)['summary'], indent=2))
        print("\n--- Cleaned CSV Data (Head) ---")
        print(pd.read_csv(csv_cleaned_path).head())
    except Exception as e:
        logger.error(f"Failed to read CSV results: {e}")

    logger.info("="*30)
    logger.info("JSONL CLEANING RESULTS (EXAMPLE 2)")
    logger.info("="*30)
    try:
        with open(jsonl_report_path, 'r') as f:
            print(json.dumps(json.load(f)['summary'], indent=2))
        print("\n--- Cleaned JSONL Data (Head) ---")
        print(pd.read_csv(jsonl_cleaned_path).head())
    except Exception as e:
        logger.error(f"Failed to read JSONL results: {e}")

    if _HAS_OPENPYXL and _EXAMPLE_EXCEL_CREATED:
        logger.info("="*30)
        logger.info("EXCEL CLEANING RESULTS (EXAMPLE 2)")
        logger.info("="*30)
        try:
            with open(excel_report_path, 'r') as f:
                print(json.dumps(json.load(f)['summary'], indent=2))
            print("\n--- Cleaned Excel Data (Head) ---")
            print(pd.read_csv(excel_cleaned_path).head())
        except Exception as e:
            logger.error(f"Failed to read Excel results: {e}")

# ---
# MAIN: interactive CLI wrapper
# ---

def interactive_menu():
    """Simple interactive menu for running the pipeline.

    Works in local Python and in Colab (if `google.colab.files` is available).
    """
    print("What would you like to do?")
    print("[1] Upload a new file from my computer (Colab only)")
    print("[2] Use a file path already on the machine")
    print("[3] Run the built-in demo examples")
    print("[4] Run analysis on ./vehicles.csv (if present)")
    print("-" * 30)

    try:
        choice = input("Enter 1, 2 or 3: ").strip()
    except Exception:
        choice = '3'

    if choice == '1':
        if files is None:
            logger.error("Upload option not available in this environment. Please use option 2.")
            return
        try:
            print("Please upload your data file (e.g., .csv, .jsonl, .xlsx, .json):")
            uploaded = files.upload()
            if not uploaded:
                logger.warning("No file uploaded.")
                return
            input_filename = list(uploaded.keys())[0]
            logger.info(f"User uploaded file: '{input_filename}'")
            run_pipeline_on_user_file(input_filename, chunksize=100000)
        except Exception as e:
            logger.error(f"An error occurred during file upload or processing: {e}")

    elif choice == '2':
        try:
            input_filename = input("Please enter the full path to your file: ").strip()
            if os.path.exists(input_filename):
                logger.info(f"Using file at path: {input_filename}")
                run_pipeline_on_user_file(input_filename, chunksize=100000)
            else:
                logger.error(f"File not found at path: {input_filename}")
        except Exception as e:
            logger.error(f"An error occurred: {e}")

    elif choice == '3':
        run_demo_examples()

    elif choice == '4':
        vehicles_path = os.path.join(os.getcwd(), "vehicles.csv")
        if os.path.exists(vehicles_path):
            logger.info(f"Found vehicles.csv at: {vehicles_path}. Running pipeline.")
            run_pipeline_on_user_file(vehicles_path, chunksize=100000)
        else:
            logger.error(f"Could not find vehicles.csv at expected path: {vehicles_path}")

    else:
        logger.warning("Invalid choice. Please run the script again and enter 1, 2, or 3.")


if __name__ == "__main__":
    interactive_menu()